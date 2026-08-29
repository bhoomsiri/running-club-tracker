"""The renderer's job is to not lose anything on the way into the file.

Read back with openpyxl rather than asserted against bytes: what matters is what a
person opening the workbook sees, and that is what reading it back measures.
"""

from __future__ import annotations

import re
import zipfile
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.adapters.export.xlsx_workbook import XlsxWorkbookRenderer
from app.application.ports.workbook_renderer import Sheet


def read(content: bytes) -> object:
    return load_workbook(BytesIO(content))


def numbers_in(content: bytes) -> list[str]:
    """The numeric literals as they are written into the file.

    xlsx stores numbers as text and Excel reads them back as doubles — there is no
    Decimal in a spreadsheet, and openpyxl's own loader returns floats. So what "no
    float" can actually mean here is checked at the only place it is decidable: the text
    the writer put in the file.
    """
    with zipfile.ZipFile(BytesIO(content)) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml").decode()
    return re.findall(r"<v>([^<]*)</v>", xml)


class TestValuesSurviveTheTrip:
    def test_no_binary_rounding_reaches_the_file(self) -> None:
        """Golden rule #6 all the way to the cell.

        A hundred small credits added up, once in Decimal and once in float — which is
        what a ledger balance is. The float pipeline writes 7.000000000000001 into a file
        someone will treat as authoritative; the Decimal one writes 7. That difference is
        the whole reason this app carries Decimals from the ledger to here without
        converting.

        (openpyxl trims most float noise on the way out, which is why this takes a
        hundred additions to show. That it needs a realistic case rather than a toy one
        is the point: the corruption arrives quietly, in a total nobody recomputes.)
        """
        credit = "0.07"
        exact = sum([Decimal(credit)] * 100, start=Decimal("0"))
        as_float = sum([float(credit)] * 100)

        decimal_file = XlsxWorkbookRenderer().render(
            [Sheet(key="l", title="แต้ม", headers=["แต้ม"], rows=[[exact]])]
        )
        float_file = XlsxWorkbookRenderer().render(
            [Sheet(key="l", title="แต้ม", headers=["แต้ม"], rows=[[as_float]])]  # type: ignore[list-item]
        )

        assert numbers_in(decimal_file) == ["7"]
        assert numbers_in(float_file) == ["7.000000000000001"]

    def test_the_stored_value_is_the_decimals_own(self) -> None:
        """openpyxl normalises trailing zeros in the stored value — 10.20 is written as
        10.2 — because the value and its presentation are separate things in xlsx. The
        value must be exact; the places a reader sees come from the format below."""
        sheet = Sheet(
            key="ledger", title="แต้ม", headers=["แต้ม"],
            rows=[[Decimal("10.20")], [Decimal("0.10")], [Decimal("5.250")]],
        )

        assert numbers_in(XlsxWorkbookRenderer().render([sheet])) == ["10.2", "0.1", "5.25"]

    def test_the_display_format_follows_the_values_own_precision(self) -> None:
        """Three places for a distance, two for points — read off the Decimal rather
        than written down again here, so the two cannot drift. This is what puts the
        trailing zero back in front of a reader after the value dropped it."""
        sheets = [
            Sheet(key="runs", title="ผลวิ่ง", headers=["ระยะ"], rows=[[Decimal("5.250")]]),
            Sheet(key="ledger", title="แต้ม", headers=["แต้ม"], rows=[[Decimal("10.20")]]),
        ]

        workbook = read(XlsxWorkbookRenderer().render(sheets))

        assert workbook["ผลวิ่ง"]["A2"].number_format == "0.000"  # type: ignore[index]
        assert workbook["แต้ม"]["A2"].number_format == "0.00"  # type: ignore[index]

    def test_an_integer_column_is_left_alone(self) -> None:
        sheet = Sheet(key="rewards", title="ของรางวัล", headers=["คงเหลือ"], rows=[[14]])

        workbook = read(XlsxWorkbookRenderer().render([sheet]))

        assert workbook["ของรางวัล"]["A2"].value == 14  # type: ignore[index]
        assert workbook["ของรางวัล"]["A2"].number_format == "General"  # type: ignore[index]

    def test_a_utc_timestamp_is_written_in_club_time(self) -> None:
        """Excel has no concept of an offset and openpyxl refuses a tz-aware datetime
        outright, so the conversion has to happen somewhere. Here, once, +07."""
        sheet = Sheet(
            key="runs", title="ผลวิ่ง", headers=["ส่งเมื่อ"],
            rows=[[datetime(2026, 8, 26, 2, 30, tzinfo=UTC)]],
        )

        workbook = read(XlsxWorkbookRenderer().render([sheet]))

        assert workbook["ผลวิ่ง"]["A2"].value == datetime(2026, 8, 26, 9, 30)  # type: ignore[index]

    def test_dates_strings_bools_and_blanks_all_arrive(self) -> None:
        sheet = Sheet(
            key="members", title="สมาชิก",
            headers=["วันที่", "ชื่อ", "เปิดใช้", "ว่าง"],
            rows=[[date(2026, 8, 20), "นักวิ่ง", True, None]],
        )

        workbook = read(XlsxWorkbookRenderer().render([sheet]))
        row = [cell.value for cell in workbook["สมาชิก"][2]]  # type: ignore[index]

        assert row == [datetime(2026, 8, 20), "นักวิ่ง", True, None]


class TestTheWorkbookItself:
    def test_one_tab_per_sheet_in_order_with_thai_titles(self) -> None:
        sheets = [
            Sheet(key="members", title="สมาชิก", headers=["ชื่อ"], rows=[]),
            Sheet(key="runs", title="ผลวิ่ง", headers=["ชื่อ"], rows=[]),
        ]

        workbook = read(XlsxWorkbookRenderer().render(sheets))

        assert workbook.sheetnames == ["สมาชิก", "ผลวิ่ง"]  # type: ignore[attr-defined]

    def test_the_default_empty_sheet_is_not_left_behind(self) -> None:
        workbook = read(
            XlsxWorkbookRenderer().render(
                [Sheet(key="members", title="สมาชิก", headers=["ชื่อ"], rows=[])]
            )
        )

        assert len(workbook.sheetnames) == 1  # type: ignore[attr-defined]

    def test_headers_are_written_and_frozen(self) -> None:
        sheet = Sheet(key="members", title="สมาชิก", headers=["ชื่อ", "หน่วยงาน"], rows=[])

        workbook = read(XlsxWorkbookRenderer().render([sheet]))

        assert [c.value for c in workbook["สมาชิก"][1]] == ["ชื่อ", "หน่วยงาน"]  # type: ignore[index]
        assert workbook["สมาชิก"].freeze_panes == "A2"  # type: ignore[index]

    def test_a_title_excel_would_reject_is_made_safe(self) -> None:
        """Excel refuses []:*?/\\ in a sheet name and truncates past 31 characters.
        A workbook that will not open is worse than a renamed tab."""
        sheet = Sheet(key="x", title="a/b[c]:" + "y" * 40, headers=["h"], rows=[])

        workbook = read(XlsxWorkbookRenderer().render([sheet]))
        name = workbook.sheetnames[0]  # type: ignore[attr-defined]

        assert len(name) <= 31
        assert not set(name) & set("[]:*?/\\")
