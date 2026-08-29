"""The export endpoint end to end, through the real repositories and a real database.

The use-case tests cover the rules with fakes. These cover the two things only a real
run can show: that what comes back is a workbook a spreadsheet program will open, and
that the audit row is really in the table afterwards rather than merely staged.
"""

from __future__ import annotations

from collections.abc import Iterator
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from uuid import UUID, uuid4

import pytest
import sqlalchemy as sa
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session, sessionmaker

from app.adapters.persistence import models
from app.api import deps
from app.config import Settings
from app.main import create_app
from tests.integration.conftest import CONSENT_VERSION, StubVerifier

pytestmark = pytest.mark.integration

ALICE, ADMIN, BOSS = "user_alice", "user_admin", "user_boss"
CAMPAIGN = UUID("cccccccc-cccc-cccc-cccc-cccccccccccc")
GRANTED_AT = datetime(2026, 8, 1, 9, 0, tzinfo=UTC)
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def club(session_factory: sessionmaker[Session]) -> dict[str, UUID]:
    """Alice — consented, with a run and a health record — plus an admin and the boss."""
    ids = {"alice": uuid4(), "admin": uuid4(), "boss": uuid4()}
    with session_factory() as session:
        session.add(
            models.Campaign(
                id=CAMPAIGN, code="hundred-km-2026", name="100 km",
                type="cumulative_distance", starts_on=date(2026, 8, 1),
                ends_on=date(2026, 12, 31), config={"target_km": 100},
            )
        )
        session.add(
            models.Member(
                id=ids["alice"], clerk_user_id=ALICE, display_name="Alice",
                full_name_th="อลิศ ใจดี", sex="female", birth_date=date(1990, 5, 20),
                department="อายุรกรรม", position="พยาบาล", shirt_size="M",
                phone="0812345678", emergency_contact_name="สมหญิง ใจดี",
                emergency_contact_phone="0898765432",
            )
        )
        session.add(
            models.Member(
                id=ids["admin"], clerk_user_id=ADMIN, display_name="Admin", role="admin"
            )
        )
        session.add(
            models.Member(
                id=ids["boss"], clerk_user_id=BOSS, display_name="Boss", role="superuser"
            )
        )
        session.flush()
        session.add(
            models.RunEntry(
                id=uuid4(), member_id=ids["alice"], distance_km=Decimal("5.250"),
                duration_seconds=1800, run_date=date(2026, 8, 20), evidence_key="k",
                evidence_sha256="b" * 64, source="app_screenshot",
            )
        )
        session.add(
            models.Consent(
                id=uuid4(), member_id=ids["alice"], purpose="health_data",
                version=CONSENT_VERSION, granted_at=GRANTED_AT,
            )
        )
        session.add(
            models.HealthRecord(
                id=uuid4(), member_id=ids["alice"], campaign_id=CAMPAIGN, phase="before",
                measured_on=date(2026, 8, 1), weight_kg=Decimal("72.50"),
                retention_until=datetime(2028, 8, 1, tzinfo=UTC),
            )
        )
        session.commit()
    return ids


def auth(who: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {who}"}


def client_with_health_export(
    session_factory: sessionmaker[Session], settings: Settings, enabled: bool
) -> Iterator[TestClient]:
    """The default `client` fixture leaves the flag off; this one can turn it on."""
    tuned = settings.model_copy(update={"health_export_enabled": enabled})
    app = create_app(tuned)
    app.dependency_overrides[deps.get_settings_dep] = lambda: tuned
    app.dependency_overrides[deps.get_session_factory_dep] = lambda: session_factory
    app.dependency_overrides[deps.get_token_verifier] = StubVerifier
    with TestClient(app) as test_client:
        yield test_client


@pytest.fixture
def client_with_health(
    session_factory: sessionmaker[Session], settings: Settings
) -> Iterator[TestClient]:
    yield from client_with_health_export(session_factory, settings, enabled=True)


def audit_rows(session_factory: sessionmaker[Session]) -> list[models.AuditLog]:
    with session_factory() as session:
        return list(session.execute(sa.select(models.AuditLog)).scalars())


class TestTheFileItself:
    def test_the_superuser_gets_a_workbook_a_spreadsheet_can_open(
        self, client: TestClient, club: dict[str, UUID]
    ) -> None:
        response = client.get("/admin/export", headers=auth(BOSS))

        assert response.status_code == 200
        assert response.headers["content-type"] == XLSX
        assert ".xlsx" in response.headers["content-disposition"]

        workbook = load_workbook(BytesIO(response.content))
        assert workbook.sheetnames == [
            "สมาชิก", "ผลวิ่ง", "การแลกรางวัล", "แต้ม", "ของรางวัล", "กิจกรรม"
        ]

    def test_the_run_arrives_with_its_distance_intact(
        self, client: TestClient, club: dict[str, UUID]
    ) -> None:
        response = client.get("/admin/export", headers=auth(BOSS))
        runs = load_workbook(BytesIO(response.content))["ผลวิ่ง"]

        assert runs["C2"].value == 5.25  # openpyxl reads numbers back as floats
        assert runs["C2"].number_format == "0.000"  # the stored precision is preserved

    def test_an_admin_is_refused(self, client: TestClient, club: dict[str, UUID]) -> None:
        assert client.get("/admin/export", headers=auth(ADMIN)).status_code == 403

    def test_an_ordinary_member_is_refused(
        self, client: TestClient, club: dict[str, UUID]
    ) -> None:
        assert client.get("/admin/export", headers=auth(ALICE)).status_code == 403


class TestTheAuditTrail:
    def test_the_export_is_recorded(
        self,
        client: TestClient,
        club: dict[str, UUID],
        session_factory: sessionmaker[Session],
    ) -> None:
        client.get("/admin/export", headers=auth(BOSS))
        rows = audit_rows(session_factory)

        assert [r.action for r in rows] == ["export_workbook"]
        assert rows[0].actor_member_id == club["boss"]
        assert rows[0].detail["members"] == 3

    def test_a_refused_export_records_nothing(
        self,
        client: TestClient,
        club: dict[str, UUID],
        session_factory: sessionmaker[Session],
    ) -> None:
        client.get("/admin/export", headers=auth(ADMIN))

        assert audit_rows(session_factory) == []


class TestTheHealthFlag:
    def test_off_the_sensitive_sheets_are_absent(
        self, client: TestClient, club: dict[str, UUID]
    ) -> None:
        response = client.get("/admin/export", headers=auth(BOSS))
        names = load_workbook(BytesIO(response.content)).sheetnames

        assert not any("อ่อนไหว" in name for name in names)

    def test_on_the_sensitive_sheets_appear_and_each_read_is_audited(
        self,
        client_with_health: TestClient,
        club: dict[str, UUID],
        session_factory: sessionmaker[Session],
    ) -> None:
        response = client_with_health.get("/admin/export", headers=auth(BOSS))
        workbook = load_workbook(BytesIO(response.content))
        rows = audit_rows(session_factory)

        assert "ข้อมูลสุขภาพ (อ่อนไหว)" in workbook.sheetnames
        # Alice consented, so her health row is in the file and her name is in the log.
        health = workbook["ข้อมูลสุขภาพ (อ่อนไหว)"]
        assert health["A2"].value == "อลิศ ใจดี"
        assert {r.action for r in rows} == {
            "export_workbook", "view_health", "view_contact"
        }
        assert {r.subject_member_id for r in rows if r.action == "view_health"} == {
            club["alice"]
        }

    def test_no_measurement_reaches_the_audit_detail(
        self,
        client_with_health: TestClient,
        club: dict[str, UUID],
        session_factory: sessionmaker[Session],
    ) -> None:
        """Golden rule #8, checked against what is actually in the table."""
        client_with_health.get("/admin/export", headers=auth(BOSS))
        stored = " ".join(str(r.detail) for r in audit_rows(session_factory))

        for forbidden in ("72.5", "0898765432", "1990-05-20"):
            assert forbidden not in stored
