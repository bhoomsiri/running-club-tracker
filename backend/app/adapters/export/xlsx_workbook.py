"""Sheets in, an .xlsx file out. The only place openpyxl is imported.

Two things it is careful about.

**No float ever touches a number on its way in.** A spreadsheet has no Decimal — xlsx
stores numbers as text and Excel reads them as doubles — so what this can guarantee is
narrower than "Decimals stay Decimals" and still worth having: the text written into the
file is the Decimal's own, never a float's. A hundred credits of 0.07 summed in Decimal
is written `7`; summed in float it is written `7.000000000000001`, into a file people
treat as authoritative (golden rule #6).

openpyxl drops trailing zeros from the stored value — 10.20 is written 10.2 — because
value and presentation are separate in xlsx. The presentation is put back by a number
format derived from the value's own exponent rather than guessed, so a balance stored to
two places shows two and a distance stored to three shows three.

**Timezones are dropped, not converted.** Excel has no concept of an offset and
openpyxl refuses a tz-aware datetime outright. Every timestamp in this app is UTC in the
database, so each is converted to the club's local time once, here, and written naive —
which is what a reader in Ratchaburi expects to see. Doing it in the use case instead
would put a display concern in the business layer.
"""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.application.ports.workbook_renderer import Cell, Sheet
from app.domain.calendar import CLUB_TIMEZONE

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Excel's own limit is 31 characters and it rejects these five outright.
MAX_TITLE = 31
_FORBIDDEN_IN_TITLE = str.maketrans({ch: "-" for ch in "[]:*?/\\"})

MIN_WIDTH, MAX_WIDTH = 10, 44


class XlsxWorkbookRenderer:
    def render(self, sheets: Sequence[Sheet]) -> bytes:
        workbook = Workbook()
        # A new Workbook comes with one empty sheet; the first real one replaces it.
        workbook.remove(workbook.active)

        for sheet in sheets:
            worksheet = workbook.create_sheet(_safe_title(sheet.title))
            worksheet.append(list(sheet.headers))
            for row in sheet.rows:
                worksheet.append([_to_cell(value) for value in row])

            for column, header in enumerate(sheet.headers, start=1):
                cell = worksheet.cell(row=1, column=column)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                worksheet.column_dimensions[get_column_letter(column)].width = _width(
                    header, [row[column - 1] for row in sheet.rows if column <= len(row)]
                )

            _apply_number_formats(worksheet, sheet)
            # Headers stay visible while someone scrolls a hundred members.
            worksheet.freeze_panes = "A2"
            if sheet.rows:
                worksheet.auto_filter.ref = worksheet.dimensions

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()


def _to_cell(value: Cell) -> object:
    if isinstance(value, datetime):
        # Excel cannot hold an offset; openpyxl raises rather than guess. UTC in, club
        # time out, naive — see the module note.
        return value.astimezone(CLUB_TIMEZONE).replace(tzinfo=None)
    return value


def _apply_number_formats(worksheet: Any, sheet: Sheet) -> None:
    """Format each numeric column to the precision its own values carry.

    Read off `Decimal.as_tuple().exponent` rather than hardcoded per column: the domain
    already decided that a distance has three places and points have two, and writing
    those numbers down again here is how the two drift apart.
    """
    for index in range(len(sheet.headers)):
        places = _decimal_places(row[index] for row in sheet.rows if index < len(row))
        if places == 0:
            continue
        column = get_column_letter(index + 1)
        for row_number in range(2, len(sheet.rows) + 2):
            worksheet[f"{column}{row_number}"].number_format = f"0.{'0' * places}"


def _decimal_places(values: Iterable[Cell]) -> int:
    """The most places any Decimal in the column carries; 0 if it holds none."""
    places = 0
    for value in values:
        if not isinstance(value, Decimal):
            continue
        exponent = value.as_tuple().exponent
        # NaN and Infinity report a string exponent; neither can reach here from the
        # domain, and neither has a number of decimal places if it did.
        if isinstance(exponent, int) and -exponent > places:
            places = -exponent
    return places


def _width(header: str, values: Sequence[Cell]) -> int:
    longest = max((len(str(v)) for v in values if v is not None), default=0)
    return min(MAX_WIDTH, max(MIN_WIDTH, len(header) + 2, longest + 2))


def _safe_title(title: str) -> str:
    return title.translate(_FORBIDDEN_IN_TITLE)[:MAX_TITLE]
